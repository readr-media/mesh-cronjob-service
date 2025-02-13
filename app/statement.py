'''
    This module provides all the function about google analytics and bigquery.
    Help you get the information to create the financial statements.
'''
import os
from google.analytics.data_v1beta import BetaAnalyticsDataClient
from google.analytics.data_v1beta.types import (
    DateRange,
    Dimension,
    Metric,
    RunReportRequest,
    Filter, 
    FilterExpression,
    FilterExpressionList
)
from google.cloud import bigquery as bq
from datetime import datetime, timezone
import math
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from dateutil.relativedelta import relativedelta
from app.gql import gql_query

homepage_title = "READr Mesh 讀選"
newpage_title  = "最新 | READr Mesh 讀選"
socialpage_title = "社群 | READr Mesh 讀選"

gql_sponsorships = '''
    query sponsorships{{
      sponsorships(where: {{status: {{equals: Success}}, createdAt: {{gt: "{START_TIME}" }} }}){{
        id
    	publisher{{
          id
        }}
        fee
      }}
    }}
'''

gql_statement_publishers = '''
query Publishers{
  publishers(where: {is_active: {equals: true}}){
    id
    title
    customId
  }
}
'''

gql_create_statements = '''
mutation createStatements($data: [StatementCreateInput!]!){
  createStatements(data: $data){
    id
  }
}
'''

gql_create_revenues = '''
mutation createRevenues($data: [RevenueCreateInput!]!){
  createRevenues(data: $data){
    id
  }
}
'''

gql_query_exchanges = '''
query exchanges{{
  exchanges(where: {{createdAt: {{gte: "{START_DATE}" }}, status: {{equals: Success}} }}, orderBy: {{id: desc}}){{
    publisher{{
      id
    }}
    tid
    exchangeVolume
    createdAt
  }}
}}
'''

gql_query_revenues = '''
query revenues{{
  revenues(where: {{createdAt: {{gte: "{START_DATE}" }}, type: {{in: [story_ad_revenue]}} }}, orderBy: {{id: desc}}){{
    publisher{{
      id
    }}
    type
    value
    start_date
  }}
}}
'''

def getRevenues(ga_resource_id, ga_months):
    # setup ga days
    current_time = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    start_datetime = current_time - relativedelta(months=ga_months)
    start_date = datetime.strftime(start_datetime, '%Y-%m-%d')
    
    # setup filter criteria
    filter_criteria = FilterExpression(
        or_group=FilterExpressionList(
            expressions=[
                FilterExpression(
                    filter=Filter(
                        field_name="pageTitle",
                        string_filter=Filter.StringFilter(
                            match_type=Filter.StringFilter.MatchType.EXACT,
                            value=homepage_title
                        )
                    )
                ),
                FilterExpression(
                    filter=Filter(
                        field_name="pageTitle",
                        string_filter=Filter.StringFilter(
                            match_type=Filter.StringFilter.MatchType.EXACT,
                            value=newpage_title
                        )
                    )
                ),
                FilterExpression(
                    filter=Filter(
                        field_name="pageTitle",
                        string_filter=Filter.StringFilter(
                            match_type=Filter.StringFilter.MatchType.EXACT,
                            value=socialpage_title
                        )
                    )
                )
            ]
        )
    )

    # send request
    request = RunReportRequest(
        property=f"properties/{ga_resource_id}",
        dimensions=[
            Dimension(name="pageTitle"),
        ],
        metrics=[
            Metric(name="totalAdRevenue"),  # 使用者
        ],
        date_ranges=[DateRange(start_date=start_date, end_date="today")],
        dimension_filter=filter_criteria,
    )
    client = BetaAnalyticsDataClient()
    response = client.run_report(request)

    # parse response
    revenue_table = {}
    total_revenue = 0
    for row in response.rows:
        dimension_value = str(row.dimension_values[0].value)
        metric_value = float(row.metric_values[0].value)
        revenue_table[dimension_value] = metric_value
        total_revenue += metric_value
    revenue_table['total'] = total_revenue
    return revenue_table

def getPublisherPageview(db_name: str, table_name: str, start_time: str):
    QUERY = (
        f'SELECT jsonPayload.complementary.targetid, COUNT(jsonPayload.complementary.targetid) AS view FROM `{db_name}.{table_name}` '
        f'WHERE (resource.type="global") AND (jsonPayload.type="click-story" OR jsonPayload.type="click-related-story") AND (jsonPayload.complementary.publishertarget="publisher") AND timestamp >= TIMESTAMP("{start_time}")'
        'GROUP BY jsonPayload.complementary.targetid;'
    )
    client = bq.Client()
    rows = client.query(QUERY).result()
    
    # parse response, pv_table contains {publisher_id: pv_count} relationship
    pv_table = {}
    for row in rows:
        pv_table[row.targetid] = row.view
    return pv_table

def calculateMutualFund(homepage_revenue: float, newpage_revenue: float):
    '''
        共同基金池資金 = 首頁廣告收益0.2+最新頁面廣告收益0.05
    '''
    return homepage_revenue*0.2+newpage_revenue*0.05

def calculatePlatformIncome(homepage_revenue: float, newpage_revenue: float, socialpage_revenue: float, collection_ad_revenue: float, story_ad_revenue: float):
    '''
        平台收益 = (首頁廣告收益0.5+最新廣告收益0.5+社群廣告收益0.5)+(集錦廣告收益0.5+文章廣告收益*0.45)
    '''
    return (homepage_revenue+newpage_revenue+socialpage_revenue)*0.5+(collection_ad_revenue*0.5+story_ad_revenue*0.45)

def publisherSponsorshipShare(gql_endpoint, mutual_fund):
    # fetch data
    current_time = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    start_datetime = (current_time - relativedelta(months=1)).isoformat().replace('+00:00', 'Z')
    data = gql_query(gql_endpoint, gql_sponsorships.format(START_TIME=start_datetime))
    sponsorships = data['sponsorships']

    # calculate statistic from sponsorships
    sponsor_table = {}
    total_fee = 0
    for sponsorship in sponsorships:
        publisher_id = sponsorship['publisher']['id']
        fee = sponsorship['fee']
        sponsor_table[publisher_id] = sponsor_table.get(publisher_id, 0)+fee
        total_fee += fee
    publisher_share_table = { pid: (fee/total_fee)*mutual_fund for pid, fee in sponsor_table.items() }
    return publisher_share_table

def createRevenuesData(gql_endpoint, shares_table: dict, start_date: str, end_date: str):
    var_revenues = {
        "data": []
    }
    for pid, data in shares_table.items():
        title, sponsorship_share, pv_share = data['title'], data['sponsorship_share'], data['pv_share']
        var_revenues["data"].append({
            "publisher": {
                "connect": {
                    "id": pid
                }
            },
            "title": f"{title}基金池分潤",
            "type": "mutual_fund_revenue",
            "value": sponsorship_share,
            "start_date": start_date,
            "end_date": end_date
        })
        var_revenues["data"].append({
            "publisher": {
                "connect": {
                    "id": pid
                }
            },
            "title": f"{title}廣告分潤",
            "type": "story_ad_revenue",
            "value": pv_share,
            "start_date": start_date,
            "end_date": end_date
        })
        print(f"pid: {pid}, sponsorship_share: {sponsorship_share}, and pv_share: {pv_share}")
    data = gql_query(gql_endpoint, gql_create_revenues, var_revenues)
    return data

def createMonthStatement(start_date: str, end_date: str, gql_endpoint: str, adsense_revenue: float, gam_revenue: float, mesh_income: float, mutual_fund: float, user_points: int, publisher_share_table: dict, pv_table, adsense_complementary: str="", gam_complementary: str="", point_complementary: str=""):
    wb = Workbook()
    ws = wb.active
    current_time = datetime.now()
    date = current_time.strftime("%Y-%m-%d")
    
    # style setting
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    precision = "{:.3f}"
    
    # for all the profit
    start_row = 1
    ws.merge_cells(f"A{start_row}:C{start_row}")
    ws[f"A{start_row}"].fill = orange_fill
    ws[f'A{start_row}'] = "收益總覽"
    ws[f'A{start_row+1}'], ws[f'B{start_row+1}'], ws[f'C{start_row+1}'] = "項目", "金額(TWD)", "備註"
    ws[f'A{start_row+2}'], ws[f'B{start_row+2}'], ws[f'C{start_row+2}'] = "Adsense收益", precision.format(adsense_revenue), adsense_complementary
    ws[f'A{start_row+3}'], ws[f'B{start_row+3}'], ws[f'C{start_row+3}'] = "GAM收益", precision.format(gam_revenue), gam_complementary
    ws[f'A{start_row+4}'], ws[f'B{start_row+4}'] = "總收益", precision.format(adsense_revenue + gam_revenue)

    # for income
    income_start_row = start_row+6
    ws.merge_cells(f"A{income_start_row}:C{income_start_row}")
    ws[f"A{income_start_row}"].fill = orange_fill
    ws[f'A{income_start_row}'] = "平台收入"
    ws[f'A{income_start_row+1}'], ws[f'B{income_start_row+1}'], ws[f'C{income_start_row+1}'] = "項目", "金額(TWD)", "備註"
    ws[f'A{income_start_row+2}'], ws[f'B{income_start_row+2}'] = "Mesh平台收入", precision.format(mesh_income)

    # for mutual fund
    fund_start_row = income_start_row+4
    ws.merge_cells(f"A{fund_start_row}:C{fund_start_row}")
    ws[f"A{fund_start_row}"].fill = orange_fill
    ws[f'A{fund_start_row}'] = "媒體共同基金池"
    ws[f'A{fund_start_row+1}'], ws[f'B{fund_start_row+1}'], ws[f'C{fund_start_row+1}'] = "項目", "金額(TWD)", "點數(MSP)"
    ws[f'A{fund_start_row+2}'], ws[f'B{fund_start_row+2}'], ws[f'C{fund_start_row+2}'] = "共同基金池", precision.format(mutual_fund), math.floor(mutual_fund)

    # user all points
    point_start_row = fund_start_row+4
    ws.merge_cells(f"A{point_start_row}:C{point_start_row}")
    ws[f"A{point_start_row}"].fill = orange_fill
    ws[f'A{point_start_row}'] = "用戶點數"
    ws[f'A{point_start_row+1}'], ws[f'B{point_start_row+1}'], ws[f'C{point_start_row+1}'] = "項目", "點數(MSP)", "備註"
    ws[f'A{point_start_row+2}'], ws[f'B{point_start_row+2}'], ws[f'C{point_start_row+2}'] = "點數總合", user_points, point_complementary

    # for publisher shares
    publisher_start_row = point_start_row+4
    ws.merge_cells(f"A{publisher_start_row}:C{publisher_start_row}")
    ws[f"A{publisher_start_row}"].fill = orange_fill
    ws[f'A{publisher_start_row}'] = "媒體廣告分潤"
    ws[f'A{publisher_start_row+1}'], ws[f'B{publisher_start_row+1}'], ws[f'C{publisher_start_row+1}'] = "媒體名稱", "共同基金池分潤(TWD)", "文章頁廣告分潤(TWD)"

    data = gql_query(gql_endpoint, gql_statement_publishers)
    publishers = data['publishers']
    total_pv = sum(pv_table.values())
    if total_pv==0:
        total_pv = 1 # avoid divide by 0 issue
    index = publisher_start_row+2
    
    shares_table = {}
    for idx, publisher in enumerate(publishers):
        id, title = publisher['id'], publisher['title']
        sponsorship_share = publisher_share_table.get(str(id), 0.0)
        pv_share = (pv_table.get(str(id), 0.0)/total_pv)*gam_revenue
        ws[f'A{index+idx}'], ws[f'B{index+idx}'], ws[f'C{index+idx}'] = title, precision.format(sponsorship_share), precision.format(pv_share)
        shares_table[id] = {
            "title": title,
            "sponsorship_share": sponsorship_share,
            "pv_share": pv_share
        }
    createRevenuesData(gql_endpoint, shares_table, start_date, end_date)
    
    # file
    folder = os.path.join("statements", "general")
    filename = os.path.join(folder, f"monthly-statement-{date}.xlsx")
    if not os.path.exists(folder):
        os.makedirs(folder)
    wb.save(filename)
    return filename


def createQuarterStatements(gql_endpoint: str, domain: str, start_date: str, end_date: str, charge_percent: float=0.1):
    current_time = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    date = current_time.strftime("%Y-%m-%d")
    filenames = []
    
    # prefetching the necessary data
    data = gql_query(gql_endpoint, gql_statement_publishers)
    publishers = data['publishers']
    
    data = gql_query(gql_endpoint, gql_query_exchanges.format(START_DATE=start_date))
    exchanges = data['exchanges']
    exchange_table = {} # mapping publisher_id to exchange records
    for exchange in exchanges:
        pid = exchange['publisher']['id']
        exchange_list = exchange_table.setdefault(pid, [])
        exchange_list.append(exchange)

    data = gql_query(gql_endpoint, gql_query_revenues.format(START_DATE=start_date))
    revenues = data['revenues']
    revenue_table = {}
    for revenue in revenues:
        pid = exchange['publisher']['id']
        revenue_list = revenue_table.setdefault(pid, [])
        revenue_list.append(revenue)
    
    # data processing
    var_statements = {
        "data": []
    }
    for publisher in publishers:
        pid, customId, title = publisher['id'], publisher['customId'], publisher['title']
        folder = os.path.join("statements", "media", customId)
        if not os.path.exists(folder):
            os.makedirs(folder)
        filename = os.path.join(folder, f"quarter-statement-{date}.xlsx")
        
        # excel: global setting
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 70
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        
        # excel: title
        start_row = 1
        ws.merge_cells(f"A{start_row}:F{start_row}")
        ws[f'A{start_row}'] = f"報表區間: {start_date}-{end_date}"
        ws[f'A{start_row+1}'], ws[f'B{start_row+1}'], ws[f'C{start_row+1}'] = "建立日期", "金流編號", "項目"
        ws[f'D{start_row+1}'], ws[f'E{start_row+1}'], ws[f'F{start_row+1}'] = "收取金額", "手續費", "實際收取金額"
        
        # excel: add exchanges information
        item_row = start_row+2 
        publisher_exchanges = exchange_table.get(pid, [])
        for exchange in publisher_exchanges:
            tid = exchange['tid']
            exchangeVolume = exchange['exchangeVolume']
            charge = math.ceil(exchangeVolume*charge_percent)
            createdAt = exchange['createdAt']
            ws[f'A{item_row}'], ws[f'B{item_row}'], ws[f'C{item_row}'] = createdAt, tid, "點數兌換"
            ws[f'D{item_row}'], ws[f'E{item_row}'], ws[f'F{item_row}'] = exchangeVolume, charge, (exchangeVolume-charge)
            item_row += 1
        publisher_revenues = revenue_table.get(pid, [])
        for revenue in publisher_revenues:
            type_name = revenue['type']
            if type_name != "story_ad_revenue":
                continue
            type_name = "廣告收益"
            start_date = revenue['start_date']
            month = datetime.strptime(start_date, '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%m')
            item_name = f"{month}月{type_name}"
            value = revenue['value']
            charge = math.ceil(value*charge_percent)
            ws[f'A{item_row}'], ws[f'B{item_row}'], ws[f'C{item_row}'] = start_date, "", item_name
            ws[f'D{item_row}'], ws[f'E{item_row}'], ws[f'F{item_row}'] = value, charge, (value-charge)
            item_row += 1
            
        # file processing
        wb.save(filename)
        filenames.append(filename)
        var_statements["data"].append({
            "title": f"{title}每期媒體報表",
            "type": "quarter",
            "url": f"{domain}{filename}",
            "publisher": {
                "connect": {
                    "id": pid
                }
            },
            "start_date": start_date,
            "end_date": end_date,
        })
        
    # update CMS
    gql_query(gql_endpoint, gql_create_statements, var_statements)
    return filenames